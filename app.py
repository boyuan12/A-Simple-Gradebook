from flask import Flask, render_template, request, flash, redirect, session, json, abort
from flask_socketio import SocketIO, emit
import sqlite3
import random
from helpers import upload_file, send_email, random_string, login_required, dcode_to_did
from werkzeug.security import generate_password_hash, check_password_hash
from termcolor import colored
from werkzeug.exceptions import HTTPException
from openpyxl import load_workbook
import os
"""
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration

sentry_sdk.init(
	dsn="https://6b4e24f141bd48a582f5b24b96095bbf@o393097.ingest.sentry.io/5241674",
	integrations=[FlaskIntegration()]
)
"""

app = Flask(__name__)
app.config["SECRET_KEY"] = "haah"
app.config['UPLOAD_FOLDER'] = "files"
app.config["DATABASE"] = "db.sqlite3"
socketio = SocketIO(app)

if not os.getenv("DATABASE_URL"):
    conn = sqlite3.connect(app.config["DATABASE"], check_same_thread=False)
    c = conn.cursor()
    BASE_URL = "127.0.0.1:5000"

else:

    from sqlalchemy import create_engine
    from sqlalchemy.orm import scoped_session, sessionmaker

    engine = create_engine(os.getenv("DATABASE_URL"))
    db = scoped_session(sessionmaker(bind=engine))
    c = db()
    conn = c
    BASE_URL = "https://a-simple-gradebook.herokuapp.com"


@app.errorhandler(HTTPException)
def handle_exception(e):
    """Return JSON instead of HTML for HTTP errors."""
    # start with the correct headers and status code from the error
    response = e.get_response()
    # replace the body with JSON
    response.data = json.dumps({
        "code": e.code,
        "name": e.name,
        "description": e.description,
    })
    if e.code == 500:
        try:
            c.execute("ROLLBACK")
            c.commit()
        except:
            pass
    response.content_type = "application/json"
    return response


def get_best_elective(wishes, exists, user_id):
    """
    ### Get the best elective based on the provided arguments.

    Args:
    - codes: list - contains subject_id that their wish electives
    - exist: list - contains periods that they already had
    - user_id: int - the primary key from users table for the student
    """

    get_elective = False
    d_id = c.execute("SELECT district_id FROM users WHERE user_id=:id", {
        "id": user_id
    }).fetchall()[0][0]
    # loop through their wishes
    for wish in wishes:
        # check for available periods
        avails = {}
        periods = []
        subs = c.execute(
            "SELECT period FROM teacher_subject WHERE subject_id=:s_id AND current_enrollment < max_enrollment",
            {
                "s_id": wish
            }).fetchall()
        for i in subs:
            periods.append(i[0])
        # compare with exists period
        if (sorted(periods) == sorted(exists)):
            continue
        else:
            avail = [x for x in subs if x not in exists]  # [3]
            period = random.choice(avail)  # choose a random period
            avails = c.execute(
                "SELECT id FROM teacher_subject WHERE subject_id=:s_id AND current_enrollment < max_enrollment AND period=:period",
                {
                    "s_id": wish,
                    "period": period[0]
                }).fetchall()
            ts_id = random.choice(avails)
            print(ts_id)
            c.execute(
                "UPDATE teacher_subject SET current_enrollment = :new WHERE id=:id",
                {
                    "new":
                    c.execute(
                        "SELECT current_enrollment FROM teacher_subject WHERE id=:id",
                        {
                            "id": ts_id[0]
                        }).fetchall()[0][0] + 1,
                    "id":
                    ts_id[0]
                })
            c.execute(
                "INSERT INTO student_subject (student_id, teacher_subject_id) VALUES (:s_id, :ts_id)",
                {
                    "s_id": user_id,
                    "ts_id": ts_id[0]
                })
            conn.commit()
            get_elective = True

    # get a random elective that is available (possible machine learning)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/create-district", methods=["GET", "POST"])
def create_school():
    if request.method == "POST":
        # check for weather user provide all required information
        if not request.form.get("email") or not request.form.get(
                "password"
        ) or not request.form.get("district_name") or not request.form.get(
                "code") or not request.form.get(
                    "address") or not request.form.get(
                        "city") or not request.form.get(
                            "state") or not request.form.get(
                                "zip") or not request.form.get("motto"):
            flash("Please make sure to fill out all required fields",
                  category="danger")
            return redirect("/create-district")

        if request.form.get("password") != request.form.get("confirmation"):
            flash("Wrong password confirmation")
            return redirect("/create-district")

        # check for whether the code already existed
        exist_code = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": request.form.get("code")
        }).fetchall()
        if len(exist_code) != 0:
            flash(
                "Your code/district abbreviation already existed. Please change your district abbreviation",
                category="danger")
            return redirect("/create-district")

        # check whether the email already existed
        exist_code = c.execute("SELECT * FROM users WHERE email=:email", {
            "email": request.form.get("email")
        }).fetchall()
        if len(exist_code) != 0:
            flash("Your email already existed. Did you already registered?",
                  category="danger")
            return redirect("/create-district")

        # get the address format
        if request.form.get("address2"):
            addr = request.form.get("address") + ", " + request.form.get(
                "address2") + " " + request.form.get(
                    "city") + ", " + request.form.get(
                        "state") + " " + request.form.get("zip")
        else:
            addr = request.form.get("address") + ", " + request.form.get(
                "city") + ", " + request.form.get(
                    "state") + " " + request.form.get("zip")

        # check for whether the address already existed
        exist_code = c.execute(
            "SELECT * FROM districts WHERE address=:address", {
                "address": addr
            }).fetchall()
        if len(exist_code) != 0:
            flash(
                "Your district address already existed. Did you already registered?",
                category="danger")
            return redirect("/create-district")

        filename = upload_file(app.config['UPLOAD_FOLDER'])

        c.execute(
            "INSERT INTO districts (name, motto, logo, address, code) VALUES (:name, :motto, :logo, :address, :code)",
            {
                "name": request.form.get("district_name"),
                "motto": request.form.get("motto"),
                "logo": filename,
                "address": addr,
                "code": request.form.get("code")
            })

        conn.commit()

        district_id = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": request.form.get("code")
        }).fetchall()[0][0]

        print(colored(district_id, "red"))

        verification_str = random_string(50)
        c.execute(
            "INSERT INTO users (school_id, name, username, password, role, district_id, email, verification, code) VALUES (0, :name, :username, :password, :role, :district_id, :email, :verification, :code)",
            {
                "name":
                request.form.get("name"),
                "username":
                request.form.get("email"),
                "password":
                generate_password_hash(request.form.get("password"),
                                       method='pbkdf2:sha256',
                                       salt_length=8),
                "role":
                "district-admin",
                "district_id":
                int(district_id),
                "email":
                request.form.get("email"),
                "verification":
                verification_str,
                "code":
                request.form.get("code") + " admin"
            })
        conn.commit()

        send_email(
            request.form.get("email"),
            "Verify Your A Simple Gradebook Account!",
            f"Please click the link to verify your account:\n{BASE_URL}/verify/{verification_str}\n. Thanks for signing up!"
        )

        return render_template(
            "success.html",
            title="Register Success - Email Verification Needed",
            details=
            "Thanks for registering! Please check your email to verify your email account. Once you click on the link in the email account, you will be directed for the next step to continue. Please check your spam folder as well. Thanks for signing up!"
        )

    else:
        return render_template("create-district.html")


@app.route("/verify/<string:token>")
def verify(token):

    results = c.execute("SELECT * FROM users WHERE verification=:verification",
                        {
                            "verification": token
                        }).fetchall()

    if len(results) != 1:
        return render_template(
            "error.html",
            title="Wrong Verification Token",
            details=
            "You provided wrong token or the email already verified. Please double check your email token. Thanks."
        )

    c.execute(
        "UPDATE users SET verification=:verify WHERE verification=:token", {
            "verify": "verify",
            "token": token
        })
    conn.commit()

    flash("Your email is successfully verified. Please login",
          category="success")

    return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":

        if not request.form.get("username") or not request.form.get(
                "password"):
            flash("Please provide all required fields", category="danger")
            return redirect("/login")

        results = c.execute("SELECT * FROM users WHERE username=:username", {
            "username": request.form.get("username")
        }).fetchall()

        if len(results) != 1:
            flash("Wrong credentials, please register before use the service",
                  category="danger")
            return redirect("/login")

        if check_password_hash(results[0][4], request.form.get("password")):
            if results[0][8] != "verify":
                flash(
                    "Please check your email for verification. You must verify your email address before you can log in",
                    category="warning")
                return redirect("/")

            session["user_id"] = results[0][0]
            session["district_code"] = c.execute(
                "SELECT code FROM districts WHERE district_id=:id", {
                    "id": results[0][6]
                }).fetchall()[0][0]
            session["district_name"] = c.execute(
                "SELECT name FROM districts WHERE district_id=:id", {
                    "id": results[0][6]
                }).fetchall()[0][0]

            if request.args.get("next"):
                return redirect(f"/{request.args.get('next')}")

            if results[0][5] == "district-admin":
                district_info = c.execute(
                    "SELECT * FROM districts WHERE district_id=:id", {
                        "id": results[0][6]
                    }).fetchall()
                session["status"] = "district-admin"
                return redirect(f"/district-admin/{district_info[0][5]}")
            elif results[0][5] == "teacher":
                session["status"] = "teacher"
                return redirect("/teacher")

        flash("Wrong credentials, please register before use the service",
              category="danger")
        return redirect("/login")

    else:
        return render_template("login.html")


@app.route("/district-admin/<string:code>")
@login_required
def district_admin_homepage(code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    district = c.execute("SELECT * FROM districts WHERE code=:code", {
        "code": code
    }).fetchall()[0][0]
    if user_info[0][5] != "district-admin" or user_info[0][6] != district:
        abort(403)

    return render_template("district-admin/dashboard.html", code=code)


@app.route("/district-admin/<string:code>/schools", methods=["GET", "POST"])
@login_required
def district_admin_dashboard_school(code):

    d_code = code

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    if user_info[0][5] != "district-admin" or user_info[0][6] != district:
        abort(403)

    if request.method == "POST":

        codes = c.execute("SELECT code FROM schools WHERE district_id=:d_id", {
            "d_id": district
        }).fetchall()

        if request.form.get("name") and request.form.get(
                "address") and request.form.get(
                    "description") and request.form.get("code"):

            # check with code already exist
            for code in codes:
                if request.form.get("code") == code[0]:
                    return render_template(
                        "error.html",
                        title="Invalid School Code",
                        details=
                        "School code already in use, please change school code.",
                        url=f"/district-admin/{d_code}/schools")

            # handle when the user submit a form manually
            c.execute(
                "INSERT INTO schools (district_id, name, address, description, code) VALUES (:district_id, :name, :address, :description, :code)",
                {
                    "district_id": int(district),
                    "name": request.form.get("name"),
                    "address": request.form.get("address"),
                    "description": request.form.get("description"),
                    "code": request.form.get("code")
                })
            conn.commit()
            return redirect(f"/district-admin/{d_code}/schools")

        elif request.files["file"]:
            # handle when user submit an excel
            filename = upload_file(app.config["UPLOAD_FOLDER"])
            wb = load_workbook(
                os.path.join(app.config["UPLOAD_FOLDER"], filename))
            sheet = wb.active

            # check if the columns is 4
            if sheet.max_column != 4:
                flash("Wrong Format: Did you have EXACTLY 4 columns?",
                      category="danger")
                return redirect(f"/district-admin/{d_code}/schools")

            for i in range(2, sheet.max_row + 1):
                name = sheet.cell(row=i, column=1).value
                address = sheet.cell(row=i, column=2).value
                description = sheet.cell(row=i, column=3).value
                code = sheet.cell(row=i, column=4).value

                c.execute(
                    "INSERT INTO schools (district_id, name, address, description, code) VALUES (:district_id, :name, :address, :description, :code)",
                    {
                        "district_id": int(district),
                        "name": name,
                        "address": address,
                        "description": description,
                        "code": code
                    })
                conn.commit()

            os.remove(os.path.join(app.config["UPLOAD_FOLDER"], filename))
            return redirect(f"/district-admin/{d_code}/schools")

    else:
        schools = c.execute(
            "SELECT * FROM schools WHERE district_id=:district_id", {
                "district_id": district
            }).fetchall()
        return render_template("district-admin/school.html",
                               schools=schools,
                               code=code)


@app.route("/district-admin/<string:d_code>/edit/school/<string:s_code>",
           methods=["GET", "POST"])
def edit_school(d_code, s_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    if request.method == "POST":
        c.execute(
            "UPDATE schools SET name=:name, address=:address, description=:description",
            {
                "name": request.form.get("name"),
                "address": request.form.get("address"),
                "description": request.form.get("description")
            })
        conn.commit()

        return redirect(f"/district-admin/{d_code}/schools")
    else:

        school_info = c.execute(
            "SELECT * FROM schools WHERE code=:code AND district_id=:d_id", {
                "code": s_code,
                "d_id": district
            }).fetchall()

        try:
            return render_template("district-admin/edit-school.html",
                                   school=school_info[0])
        except IndexError:
            return render_template(
                "error.html",
                title="School Code not found",
                details=
                "School codes not found. Please double check your school code.",
                url=f"/district-admin/{d_code}/schools")


@app.route("/district-admin/<string:d_code>/delete/school/<string:s_code>",
           methods=["GET", "POST"])
def delete_school(d_code, s_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    if request.method == "POST":

        school_id = c.execute("SELECT school_id FROM schools WHERE code=:code",
                              {
                                  "code": s_code
                              }).fetchall()[0][0]

        c.execute(
            "DELETE FROM schools WHERE code=:code and district_id=:d_code", {
                "code": s_code,
                "d_code": district
            })

        c.execute(
            "DELETE FROM users WHERE school_id=:code and district_id=:d_code",
            {
                "code": school_id,
                "d_code": district
            })

        conn.commit()
        return redirect(f"/district-admin/{d_code}/schools")

    else:
        school_info = c.execute(
            "SELECT * FROM schools WHERE code=:code AND district_id=:d_id", {
                "code": s_code,
                "d_id": district
            }).fetchall()

        try:
            return render_template("district-admin/confirm-delete-school.html",
                                   school=school_info[0],
                                   d_code=d_code,
                                   s_code=s_code)
        except IndexError:
            return render_template(
                "error.html",
                title="School Code not found",
                details=
                "School codes not found. Please double check your school code.",
                url=f"/district-admin/{d_code}/schools")


@app.route("/district-admin/<string:d_code>/teachers", methods=["GET", "POST"])
def admin_teachers_dashboard(d_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    d_id = c.execute("SELECT district_id FROM districts WHERE code=:d_code", {
        "d_code": d_code
    }).fetchall()[0][0]

    if request.method == "POST":

        if request.form.get("name") and request.form.get(
                "address") and request.form.get("role") and request.form.get(
                    "email") and request.form.get(
                        "s_code") and request.form.get(
                            "t_code") and request.form.get("subjects"):

            # get district id (NO ERROR SHOULD RAISE)
            # get school id (ERROR MAY RAISE)

            try:
                s_id = c.execute(
                    "SELECT school_id FROM schools WHERE code=:code", {
                        "code": request.form.get("s_code")
                    }).fetchall()[0][0]
            except IndexError:
                return render_template(
                    "error.html",
                    title="No School Found",
                    detail=
                    "System didn't found the school code that you substitute in. Please double check.",
                    url=f"/district-admin/{d_code}")

            pwd = random_string(10)
            ver = random_string(70)

            subjects = request.form.get("subjects").split(", ")

            results = c.execute(
                "SELECT email FROM users WHERE email=:email AND district_id=:d_id AND role='teacher'",
                {
                    "email": request.form.get("email"),
                    "d_id": d_id
                }).fetchall()

            if (request.form.get("email"), ) in results:
                flash("Error, email already exist")
                return redirect(f"/district-admin/{d_code}/teachers")

            c.execute(
                "INSERT INTO users (school_id, name, username, password, role, district_id, email, verification, address, role_description, code) VALUES (:school_id, :name, :username, :password, :role, :district_id, :email, :verification, :address, :role_description, :code)",
                {
                    "school_id": s_id,
                    "name": request.form.get("name"),
                    "username": request.form.get("email"),
                    "password": generate_password_hash(pwd),
                    "role": "teacher",
                    "district_id": d_id,
                    "email": request.form.get("email"),
                    "verification": ver,
                    "address": request.form.get("address"),
                    "role_description": request.form.get("role"),
                    "code": request.form.get("t_code")
                })

            conn.commit()

            t_id = c.execute(
                "SELECT user_id FROM users WHERE email=:email AND role='teacher' AND district_id=:d_id",
                {
                    "email": request.form.get("email"),
                    "d_id": d_id
                }).fetchall()[0][0]

            for subject in subjects:

                sub = subject.split("-")

                try:
                    sub_id = c.execute(
                        "SELECT course_id FROM courses WHERE district_id=:d_id AND code=:code",
                        {
                            "d_id": d_id,
                            "code": sub[1]
                        }).fetchall()[0][0]
                except IndexError:
                    return render_template(
                        "error.html",
                        title="Course Not Found",
                        details="Please check your course code")

                c.execute(
                    "INSERT INTO teacher_subject (teacher_id, period, subject_id, current_enrollment, max_enrollment, district_id) VALUES (:t_id, :p, :s, 0, :max, :d_id)",
                    {
                        "t_id": t_id,
                        "p": sub[0],
                        "s": sub_id,
                        "max": sub[2],
                        "d_id": d_id
                    })

                conn.commit()

            send_email(
                request.form.get("email"),
                f"Invite for joining {d_code} as a {request.form.get('role')}",
                f"If you do want to join, please go ahead and click this link: https://{BASE_URL}/verify/{ver}. Your username is {request.form.get('email')}, and your password is {pwd}."
            )

            return redirect(f"/district-admin/{d_code}/teachers")

        if request.files["file"]:

            filename = upload_file(app.config["UPLOAD_FOLDER"])
            wb = load_workbook(
                os.path.join(app.config["UPLOAD_FOLDER"], filename))
            sheet = wb.active

            if sheet.max_column != 6:
                return render_template(
                    "error.html",
                    title="Is Your Excel file correct?",
                    detail=
                    "Your excel file is not in correct format. Did you have 7 column and are using the template layout that is given?"
                )

            for i in range(2, sheet.max_row + 1):
                name = sheet.cell(row=i, column=1).value
                address = sheet.cell(row=i, column=2).value
                role = sheet.cell(row=i, column=3).value
                email = sheet.cell(row=i, column=4).value
                s_code = sheet.cell(row=i, column=5).value
                t_code = sheet.cell(row=i, column=6).value

                # check whether the teacher code already exist
                t_codes = c.execute(
                    "SELECT * FROM users WHERE district_id=:d_id AND code=:code",
                    {
                        "d_id": d_id,
                        "code": t_code
                    }).fetchall()

                if len(t_codes) != 0:
                    return render_template(
                        "error.html",
                        title="Invalid Teacher Code",
                        details=
                        "Teacher code is either invalid or already exist. Please change it. Thanks."
                    )

                try:
                    s_id = c.execute(
                        "SELECT school_id FROM schools WHERE code=:code AND district_id=:d_id",
                        {
                            "code": s_code,
                            "d_id": district
                        }).fetchall()[0][0]
                except IndexError:
                    return render_template(
                        "error.html",
                        title="No School Found",
                        details=
                        "System didn't found the school code that you substitute in. Please double check.",
                        url=f"/district-admin/{d_code}")

                pwd = random_string(10)
                ver = random_string(70)

                c.execute(
                    "INSERT INTO users (school_id, name, username, password, role, district_id, email, verification, address, role_description, code) VALUES (:school_id, :name, :username, :password, :role, :district_id, :email, :verification, :address, :role_description, :code)",
                    {
                        "school_id": s_id,
                        "name": name,
                        "username": email,
                        "password": generate_password_hash(pwd),
                        "role": "teacher",
                        "district_id": d_id,
                        "email": email,
                        "address": address,
                        "verification": ver,
                        "role_description": role,
                        "code": t_code
                    })

                conn.commit()

                send_email(
                    email, f"Invite for joining {d_code} as a {role}",
                    f"If you do want to join, please go ahead and click this link: https://{BASE_URL}/verify/{ver}. Your username is {email}, and your password is {pwd}."
                )

            os.remove(os.path.join(app.config["UPLOAD_FOLDER"], filename))

            return redirect(f"/district-admin/{d_code}")

    else:
        results = []
        teachers = c.execute(
            "SELECT * FROM users WHERE role='teacher' AND district_id=:d_id", {
                "d_id": district
            }).fetchall()

        for teacher in teachers:
            name = teacher[2]
            address = teacher[9]
            role = teacher[10]
            email = teacher[3]
            s_id = teacher[1]
            s_code = c.execute(
                "SELECT code FROM schools WHERE school_id=:s_id", {
                    "s_id": s_id
                }).fetchall()[0][0]
            t_code = teacher[11]
            result = [name, address, role, email, s_code, t_code]
            results.append(result)

        return render_template("district-admin/teachers.html",
                               teachers=results)


@app.route("/district-admin/<string:d_code>/edit/teacher/<string:t_code>",
           methods=["GET", "POST"])
def edit_teacher(d_code, t_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    d_id = c.execute("SELECT district_id FROM districts WHERE code=:code", {
        "code": d_code
    }).fetchall()[0][0]

    # check whether the teacher code is valid
    try:
        t_id = c.execute(
            "SELECT * FROM users WHERE role='teacher' AND district_id=:d_id AND code=:t_code",
            {
                "d_id": d_id,
                "t_code": t_code
            }).fetchall()[0][0]
    except IndexError:
        return render_template(
            "error.html",
            title="No Teacher Code Found",
            details="No teacher code, please double check. Thanks")

    if request.method == "POST":
        try:
            s_id = c.execute("SELECT school_id FROM schools WHERE code=:code",
                             {
                                 "code": request.form.get("s_code")
                             }).fetchall()[0][0]
        except IndexError:
            return render_template(
                "error.html",
                title="No School Found",
                detail=
                "System didn't found the school code that you substitute in. Please double check.",
                url=f"/district-admin/{d_code}")

        c.execute(
            "UPDATE users SET name=:name, address=:address, role_description=:role_description, email=:email, school_id=:school_id WHERE role='teacher' AND district_id=:district_id AND code=:code",
            {
                "name": request.form.get("name"),
                "address": request.form.get("address"),
                "role_description": request.form.get("role"),
                "email": request.form.get("email"),
                "school_id": s_id,
                "district_id": d_id,
                "code": t_code
            })

        conn.commit()

        return redirect(f"/district-admin/{d_code}/teachers")

    else:
        teacher = c.execute(
            "SELECT * FROM users WHERE role='teacher' AND district_id=:d_id AND code=:t_code",
            {
                "d_id": d_id,
                "t_code": t_code
            }).fetchall()[0]
        print(teacher)
        name = teacher[2]
        address = teacher[9]
        role = teacher[10]
        email = teacher[3]
        s_id = teacher[1]
        s_code = c.execute("SELECT code FROM schools WHERE school_id=:s_id", {
            "s_id": s_id
        }).fetchall()[0][0]
        t_code = teacher[11]
        result = [name, address, role, email, s_code, t_code]

        return render_template("district-admin/edit-teacher.html",
                               teacher=result)


@app.route("/district-admin/<string:d_code>/delete/teacher/<string:t_code>",
           methods=["GET", "POST"])
def delete_teacher(d_code, t_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    if request.method == "POST":
        c.execute(
            "DELETE FROM users WHERE district_id = :d_id AND code=:code ", {
                "d_id": district,
                "code": t_code
            })
        conn.commit()
        return redirect(f"/district-admin/{d_code}/teachers")

    else:
        try:
            name = c.execute(
                "SELECT name FROM users WHERE role='teacher' AND code=:t_code AND district_id=:d_id",
                {
                    "t_code": t_code,
                    "d_id": district
                }).fetchall()[0][0]
        except:
            return render_template(
                "error.html",
                title="No Teacher Found",
                details="No teacher code found, please double check. Thanks.")
        return render_template("district-admin/confirm-delete-teacher.html",
                               name=name,
                               d_code=d_code,
                               t_code=t_code)


@app.route("/district-admin/<string:d_code>/students", methods=["GET", "POST"])
def students(d_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)
    d_id = c.execute("SELECT * FROM districts WHERE code=:d_code", {
        "d_code": d_code
    }).fetchall()[0][0]

    if request.method == "POST":
        if request.form.get("name") and request.form.get(
                "address") and request.form.get("grade") and request.form.get(
                    "email") and request.form.get(
                        "s_code") and request.form.get("t_code"):
            """
                            d_id = dcode_to_did(d_code)
                            subjects = request.form.get("subjects").split(", ")
                            for subject in subjects:

                                sub = subject.split("-")

                                try:
                                    sub_id = c.execute(
                                        "SELECT course_id FROM courses WHERE district_id=:d_id AND code=:code",
                                        {
                                            "d_id": d_id,
                                            "code": sub[1]
                                        }).fetchall()[0][0]
                                except IndexError:
                                    return render_template(
                                        "error.html",
                                        title="Course Not Found",
                                        details="Please check your course code")

                                c.execute(
                                    "INSERT INTO teacher_subject (teacher_id, period, subject_id) VALUES (:t_id, :p, :s)",
                                    {
                                        "t_id": t_id,
                                        "p": sub[0],
                                        "s": sub_id
                                    })

                                conn.commit()
                            """
            school_id = c.execute(
                "SELECT school_id FROM schools WHERE district_id=:d_id AND code=:code",
                {
                    "d_id": d_id,
                    "code": request.form.get("s_code")
                }).fetchall()
            if len(school_id) == 0:
                return "school not found"
            c.execute(
                "INSERT INTO users (school_id, name, username, password, role, district_id, email, verification, address, role_description, code) VALUES (:s_id, :name, :user, :password, :role, :district_id, :email, :ver, :address, :role_d, :code)",
                {
                    "s_id":
                    school_id[0][0],
                    "name":
                    request.form.get("name"),
                    "user":
                    request.form.get("email"),
                    "password":
                    generate_password_hash(random_string(10),
                                           method="pbkdf2:sha256",
                                           salt_length=8),
                    "role":
                    "student",
                    "district_id":
                    d_id,
                    "email":
                    request.form.get("email"),
                    "ver":
                    "verify",
                    "address":
                    request.form.get("address"),
                    "role_d":
                    f"Grade {request.form.get('grade')} student",
                    "code":
                    request.form.get("t_code")
                })

            conn.commit()

            return redirect(f"/district-admin/{d_code}/students")
    else:
        students = c.execute(
            "SELECT * FROM users WHERE district_id=:d_id AND role='student'", {
                "d_id": d_id
            }).fetchall()
        students_list = []
        for student in students:
            school = c.execute(
                "SELECT name FROM schools WHERE school_id=:s_id", {
                    "s_id": student[1]
                }).fetchall()[0][0]
            info = [
                student[2], student[9], student[10], student[7], school,
                student[11]
            ]
            students_list.append(info)
        return render_template("district-admin/students.html",
                               students=students_list)


@app.route("/district-admin/<string:d_code>/detail/student/<string:s_code>")
@login_required
def student_detail_admin(d_code, s_code):
    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    if request.method == "POST":
        pass
    else:

        d_id = c.execute(
            "SELECT district_id FROM districts WHERE code=:d_code", {
                "d_code": d_code
            }).fetchall()[0][0]

        info = c.execute(
            "SELECT * FROM users WHERE code=:s_code AND district_id=:d_id", {
                "s_code": s_code,
                "d_id": d_id
            }).fetchall()

        s_name = c.execute(
            "SELECT name FROM schools WHERE district_id=:d_id AND school_id=:s_id",
            {
                "d_id": d_id,
                "s_id": info[0][1]
            }).fetchall()
        # print(info)
        print(info[0][0])
        ts = c.execute(
            "SELECT * FROM student_subject  JOIN teacher_subject ON teacher_subject.id=student_subject.teacher_subject_id JOIN courses ON subject_id=courses.course_id WHERE student_id=:s_id;",
            {
                "s_id": info[0][0]
            }).fetchall()

        print(ts)
        # return str(info + ts) index # 4 - current, 5 - max
        return render_template("district-admin/student-detail.html",
                               info=info,
                               ts=ts,
                               s_name=s_name)


@app.route("/district-admin/<string:d_code>/details/teacher/<string:t_code>",
           methods=["GET", "POST"])
@login_required
def district_admin_teacher_detail(d_code, t_code):

    if request.method == "POST":
        d_id = c.execute("SELECT * FROM districts WHERE code=:d_code", {
            "d_code": d_code
        }).fetchall()[0][0]
        c.execute(
            "UPDATE users SET name=:name, address=:address, email=:email, password=:password WHERE district_id=:d_id AND code=:code",
            {
                "name":
                request.form.get("name"),
                "address":
                request.form.get("address"),
                "email":
                request.form.get("email"),
                "password":
                generate_password_hash(request.form.get("password"),
                                       method="pbkdf2:sha256",
                                       salt_length=8),
                "d_id":
                d_id,
                "code":
                t_code
            })
        conn.commit()
        return render_template("success.html",
                               title="Updated Teacher Successfully!")
    else:
        user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
            "user_id": session.get("user_id")
        }).fetchall()
        try:
            district = c.execute("SELECT * FROM districts WHERE code=:code", {
                "code": d_code
            }).fetchall()[0][0]
        except IndexError:
            abort(403)

        d_id = c.execute(
            "SELECT district_id FROM districts WHERE code=:d_code", {
                "d_code": d_code
            }).fetchall()[0][0]

        info = c.execute(
            "SELECT * FROM users WHERE code=:t_code AND district_id=:d_id", {
                "t_code": t_code,
                "d_id": d_id
            }).fetchall()

        s_name = c.execute(
            "SELECT name FROM schools WHERE district_id=:d_id AND school_id=:s_id",
            {
                "d_id": d_id,
                "s_id": info[0][1]
            }).fetchall()
        # print(info)

        ts = c.execute(
            "SELECT * FROM teacher_subject JOIN courses ON courses.course_id=teacher_subject.subject_id WHERE teacher_id=:t_id ORDER BY teacher_subject.period ASC",
            {
                "t_id": info[0][0]
            }).fetchall()

        print(ts)
        # return str(info + ts) index # 4 - current, 5 - max
        return render_template("district-admin/teacher-detail.html",
                               info=info,
                               ts=ts,
                               s_name=s_name)


@app.route("/district-admin/<string:d_code>/courses", methods=["GET", "POST"])
@login_required
def district_admin_courses(d_code):

    user_info = c.execute("SELECT * FROM users WHERE user_id=:user_id", {
        "user_id": session.get("user_id")
    }).fetchall()
    try:
        district = c.execute("SELECT * FROM districts WHERE code=:code", {
            "code": d_code
        }).fetchall()[0][0]
    except IndexError:
        abort(403)

    d_id = c.execute("SELECT district_id FROM districts WHERE code=:code", {
        "code": d_code
    }).fetchall()[0][0]

    if request.method == "POST":
        courses = c.execute("SELECT code FROM courses WHERE district_id=:d_id",
                            {
                                "d_id": d_id
                            }).fetchall()
        if (request.form.get("c_code"), ) in courses:
            return render_template("error.html",
                                   title="Course code already exist",
                                   details="Please change your course code")

        # if it does pass

        s_code = c.execute(
            "SELECT * FROM schools WHERE code=:code AND district_id=:d_id", {
                "code": request.form.get("s_code"),
                "d_id": d_id
            }).fetchall()

        if len(s_code) != 1:
            return render_template("error.html",
                                   title="School code not found",
                                   details="Please check your school code")

        c.execute(
            "INSERT INTO courses (title, grade, district_id, school_id, description, code) VALUES (:title, :grade, :district_id, :school_id, :description, :code)",
            {
                "title": request.form.get("title"),
                "grade": request.form.get("grade"),
                "district_id": d_id,
                "school_id": s_code[0][0],
                "description": request.form.get("description"),
                "code": request.form.get("c_code"),
            })
        conn.commit()

        return redirect(f"/district-admin/{d_code}/courses")

    else:
        courses = c.execute("SELECT * FROM courses WHERE district_id=:d_id", {
            "d_id": d_id
        }).fetchall()
        c_list = []
        for course in courses:
            s_code = c.execute(
                "SELECT code FROM schools WHERE school_id=:s_id", {
                    "s_id": course[4]
                }).fetchall()[0][0]
            c_list.append([course[1], course[5], course[2], s_code, course[6]])
        # return str(courses)
        return render_template("district-admin/courses.html", courses=c_list)


@app.route("/district-admin/<string:d_code>/edit/course")
def edit_course():
    pass


@app.route("/email", methods=["GET", "POST"])
@login_required
def email():

    d_id = c.execute("SELECT district_id FROM districts WHERE code=:code", {
        "code": session.get("district_code")
    }).fetchall()[0][0]

    if request.method == "POST":

        if not request.form.get("receiver") or not request.form.get(
                "subject") or not request.form.get("content"):
            return render_template(
                "error.html",
                title="Please fill out all required fields",
                details=
                "Please make sure you fill out required fields when sending a email"
            )

        for i in request.form.get("receiver").split(", "):

            z = i.split(" - ")

            try:
                user_id = c.execute(
                    "SELECT user_id FROM users WHERE code=:code AND district_id=:d_id",
                    {
                        "code": z[1],
                        "d_id": d_id
                    }).fetchall()[0][0]
            except IndexError:
                if z[0] != "":
                    return render_template('error.html',
                                           title="person didn't find",
                                           details=str(z))
                else:
                    break

            c.execute(
                "INSERT INTO emails (sender_id, receiver, subject, contents, district_id) VALUES (:sender_id, :receiver, :subject, :content, :d_id)",
                {
                    "sender_id": session.get("user_id"),
                    "receiver": user_id,
                    "subject": request.form.get("subject"),
                    "content": request.form.get("content"),
                    "d_id": d_id
                })

            conn.commit()

            print(1)

        return redirect("/email")

    else:

        users = c.execute(
            "SELECT name, code FROM users WHERE district_id=:d_id", {
                "d_id":
                c.execute("SELECT district_id FROM districts WHERE code=:code",
                          {
                              "code": session.get("district_code")
                          }).fetchall()[0][0]
            }).fetchall()
        users_str = ""

        for i in range(len(users)):
            if i == len(users) - 1:
                users_str += f"{users[i][0]} - {users[i][1]}"
            else:
                users_str += f"{users[i][0]} - {users[i][1]}, "

        emails = c.execute("SELECT * FROM emails WHERE receiver=:r_id", {
            "r_id": str(session.get("user_id"))
        }).fetchall()

        if len(emails) == 0:
            return render_template("district-admin/email.html",
                                   users_str=users_str,
                                   emails=emails)

        emails_list = []

        for email in emails:

            subject = email[5]
            content = email[6]
            sender_name = c.execute(
                "SELECT name FROM users WHERE user_id=:u_id", {
                    "u_id": email[1]
                }).fetchall()[0][0]
            emails_list.append([sender_name, subject, content])

        return render_template("district-admin/email.html",
                               users_str=users_str,
                               emails=emails_list)


@app.route("/teacher")
@login_required
def teacher():
    return render_template("teacher/dashboard.html")


@app.route("/profile")
@login_required
def profile():
    user = c.execute("SELECT * FROM users WHERE user_id=:u_id", {
        "u_id": session.get("user_id")
    }).fetchall()[0]
    return render_template("district-admin/profile.html", user=user)


@app.route("/logout")
@login_required
def logout():
    session.clear()
    return redirect("/")


@app.route("/chat")
@login_required
def chat():
    return render_template("chat.html")


@socketio.on('broadcast message')
def messageDisplay(data):
    name = c.execute("SELECT name FROM users WHERE user_id=:u_id", {
        "u_id": session.get("user_id")
    }).fetchall()[0][0]
    emit("show message",
         dict(message=data["message"], name=name, timestamp=data["timestamp"]),
         broadcast=True)


@app.route("/district-admin/<string:d_code>/schedules",
           methods=["GET", "POST"])
def schedules(d_code):

    d_id = c.execute("SELECT district_id FROM districts WHERE code=:code", {
        "code": session.get("district_code")
    }).fetchall()[0][0]

    if request.method == "POST":

        if request.files["file"]:

            filename = upload_file(app.config["UPLOAD_FOLDER"])
            wb = load_workbook(
                os.path.join(app.config["UPLOAD_FOLDER"], filename))
            sheet = wb.active

            if sheet.max_column > 8:
                return render_template(
                    "error.html",
                    title="Error: Excel",
                    details=
                    "Your excel file can only contains maximum 8 columns")

            # loop through each row
            for i in range(2, sheet.max_row + 1):
                # validate student id
                student_id = c.execute(
                    "SELECT user_id FROM users WHERE code=:code", {
                        "code": sheet.cell(row=i, column=1).value
                    }).fetchall()
                if len(student_id) == 0:
                    return "Student Code is NOT VALID"
                # get elective wishes course id and store in a list
                wishes = [
                    sheet.cell(row=i, column=j).value for j in range(2, 8)
                ]
                subs = []
                for wish in wishes:
                    try:
                        sub_id = c.execute(
                            "SELECT course_id FROM courses WHERE code=:code AND district_id=:d_id",
                            {
                                "code": wish,
                                "d_id": d_id
                            }).fetchall()[0][0]
                    except IndexError:
                        return render_template("error.html",
                                               title="Course not found")
                    subs.append(sub_id)
                # get already enrolled course's period and store in a list
                ts_ids = c.execute(
                    "SELECT teacher_subject_id FROM student_subject WHERE student_id=:s_id",
                    {
                        "s_id": student_id[0][0]
                    }).fetchall()
                periods = []
                for i in ts_ids:
                    period = c.execute(
                        "SELECT period FROM teacher_subject WHERE id=:id", {
                            "id": i
                        }).fetchall()[0][0]
                    periods.append(period)
                get_best_elective(subs, periods, student_id[0][0])

            return render_template(
                "success.html",
                title="Success! Student's schedule been uploaded successfully."
            )

        return "please fill out all required fields"

    else:
        return render_template("district-admin/schedules.html")


if __name__ == "__main__":
    socketio.run(app, debug=True)